from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify, send_from_directory, session
from result_management.models.models import db, Session, Student, Subject, Mark, CourseRegistration, User
from result_management.config import Config
import os
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from datetime import datetime
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
import zipfile
import io
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from reportlab.lib.enums import TA_CENTER

app = Flask(__name__)
app.config.from_object(Config)
db.init_app(app)

with app.app_context():
    db.create_all()

def calculate_grade(total_marks, is_retake=False):
    if total_marks >= 80:
        grade_point = 4.0
        grade_letter = 'A+'
    elif total_marks >= 75:
        grade_point = 3.75
        grade_letter = 'A'
    elif total_marks >= 70:
        grade_point = 3.5
        grade_letter = 'A-'
    elif total_marks >= 65:
        grade_point = 3.25
        grade_letter = 'B+'
    elif total_marks >= 60:
        grade_point = 3.0
        grade_letter = 'B'
    elif total_marks >= 55:
        grade_point = 2.75
        grade_letter = 'B-'
    elif total_marks >= 50:
        grade_point = 2.5
        grade_letter = 'C+'
    elif total_marks >= 45:
        grade_point = 2.25
        grade_letter = 'C'
    elif total_marks >= 40:
        grade_point = 2.0
        grade_letter = 'D'
    else:
        grade_point = 0.0
        grade_letter = 'F'
    
    if is_retake and grade_letter != 'F':
        if grade_letter == 'A+':
            grade_letter = 'A'
            grade_point = 3.75
        elif grade_letter == 'A':
            grade_letter = 'A-'
            grade_point = 3.5
        elif grade_letter == 'A-':
            grade_letter = 'B+'
            grade_point = 3.25
        elif grade_letter == 'B+':
            grade_letter = 'B'
            grade_point = 3.0
        elif grade_letter == 'B':
            grade_letter = 'B-'
            grade_point = 2.75
        elif grade_letter == 'B-':
            grade_letter = 'C+'
            grade_point = 2.5
        elif grade_letter == 'C+':
            grade_letter = 'C'
            grade_point = 2.25
        elif grade_letter == 'C':
            grade_letter = 'D'
            grade_point = 2.0
    
    return grade_point, grade_letter

def convert_to_roman(num):
    val = [
        1000, 900, 500, 400,
        100, 90, 50, 40,
        10, 9, 5, 4,
        1
    ]
    syb = [
        "M", "CM", "D", "CD",
        "C", "XC", "L", "XL",
        "X", "IX", "V", "IV",
        "I"
    ]
    roman_num = ''
    i = 0
    while num > 0:
        for _ in range(num // val[i]):
            roman_num += syb[i]
            num -= val[i]
        i += 1
    return roman_num

@app.route('/')
def index():
    sessions = Session.query.all()
    return render_template('index.html', sessions=sessions)

@app.route('/add_session', methods=['GET', 'POST'])
def add_session():
    if request.method == 'POST':
        name = request.form['name']
        term = request.form['term']
        # Convert term number to Roman numeral
        term_roman = convert_to_roman(int(term))
        session = Session(name=name, term=term_roman)
        db.session.add(session)
        db.session.commit()
        flash('Session added successfully!', 'success')
        return redirect(url_for('index'))
    return render_template('add_session.html')

@app.route('/add_student/<int:session_id>', methods=['GET', 'POST'])
def add_student(session_id):
    if request.method == 'POST':
        if 'excel_file' in request.files and request.files['excel_file'].filename != '':
            # Bulk upload
            file = request.files['excel_file']
            filename = secure_filename(file.filename)
            filepath = os.path.join('uploads', filename)
            os.makedirs('uploads', exist_ok=True)
            file.save(filepath)
            wb = load_workbook(filepath)
            ws = wb.active
            added = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                student_id, name = row
                if student_id and name:
                    exists = Student.query.filter_by(student_id=str(student_id), session_id=session_id).first()
                    if not exists:
                        student = Student(student_id=str(student_id), name=name, session_id=session_id)
                        db.session.add(student)
                        added += 1
            db.session.commit()
            flash(f'{added} students added from Excel file!', 'success')
            os.remove(filepath)
            return redirect(url_for('add_student', session_id=session_id))
        else:
            # Single add
            student_id = request.form['student_id']
            name = request.form['name']
            exists = Student.query.filter_by(student_id=student_id, session_id=session_id).first()
            if exists:
                flash('Student with this ID already exists in this session.', 'danger')
            else:
                student = Student(student_id=student_id, name=name, session_id=session_id)
                db.session.add(student)
                db.session.commit()
                flash('Student added successfully!', 'success')
            return redirect(url_for('add_student', session_id=session_id))
    students = Student.query.filter_by(session_id=session_id).all()
    return render_template('add_student.html', session_id=session_id, students=students)

@app.route('/edit_student/<int:student_id>', methods=['GET', 'POST'])
def edit_student(student_id):
    student = Student.query.get_or_404(student_id)
    if request.method == 'POST':
        student.student_id = request.form['student_id']
        student.name = request.form['name']
        db.session.commit()
        flash('Student updated successfully!', 'success')
        return redirect(url_for('add_student', session_id=student.session_id))
    return render_template('edit_student.html', student=student)

@app.route('/delete_student/<int:student_id>', methods=['POST'])
def delete_student(student_id):
    student = Student.query.get_or_404(student_id)
    session_id = student.session_id
    # Delete related marks and course registrations first
    Mark.query.filter_by(student_id=student.id).delete()
    CourseRegistration.query.filter_by(student_id=student.id).delete()
    db.session.delete(student)
    db.session.commit()
    flash('Student deleted successfully!', 'success')
    return redirect(url_for('add_student', session_id=session_id))

@app.route('/add_subject/<int:session_id>', methods=['GET', 'POST'])
def add_subject(session_id):
    if request.method == 'POST':
        code = request.form['code']
        name = request.form['name']
        credit = float(request.form['credit'])
        subject_type = request.form['subject_type']
        has_retake = 'has_retake' in request.form
        
        dissertation_type = None
        if subject_type == 'Dissertation':
            dissertation_type = request.form['dissertation_type']
        
        subject = Subject(
            code=code,
            name=name,
            credit=credit,
            subject_type=subject_type,
            dissertation_type=dissertation_type,
            has_retake=has_retake,
            session_id=session_id
        )
        db.session.add(subject)
        db.session.commit()
        flash('Subject added successfully!', 'success')
        return redirect(url_for('add_subject', session_id=session_id))
    subjects = Subject.query.filter_by(session_id=session_id).all()
    return render_template('add_subject.html', session_id=session_id, subjects=subjects)

@app.route('/delete_subject/<int:subject_id>', methods=['POST'])
def delete_subject(subject_id):
    subject = Subject.query.get_or_404(subject_id)
    session_id = subject.session_id
    db.session.delete(subject)
    db.session.commit()
    flash('Subject deleted successfully!', 'success')
    return redirect(url_for('add_subject', session_id=session_id))

@app.route('/get_registered_subjects/<int:student_id>')
def get_registered_subjects(student_id):
    regs = CourseRegistration.query.filter_by(student_id=student_id).all()
    subject_ids = [reg.subject_id for reg in regs]
    subjects = Subject.query.filter(Subject.id.in_(subject_ids)).all()
    return jsonify([
        {
            'id': subject.id,
            'code': subject.code,
            'name': subject.name,
            'subject_type': subject.subject_type,
            'dissertation_type': subject.dissertation_type
        } for subject in subjects
    ])

@app.route('/add_marks/<int:session_id>', methods=['GET', 'POST'])
def add_marks(session_id):
    students = Student.query.filter_by(session_id=session_id).all()
    subjects = Subject.query.filter_by(session_id=session_id).all()
    selected_subject_id = request.args.get('subject_id', type=int)
    selected_subject = Subject.query.get(selected_subject_id) if selected_subject_id else None
    registered_students = []
    is_retake_map = {}
    if selected_subject:
        regs = CourseRegistration.query.filter_by(subject_id=selected_subject_id).all()
        student_ids = [reg.student_id for reg in regs]
        registered_students = Student.query.filter(Student.id.in_(student_ids)).all()
        for student in registered_students:
            is_retake_map[student.id] = any(r.is_retake for r in regs if r.student_id == student.id)
    if request.method == 'POST' and selected_subject:
        print(f"[DEBUG] Add Marks for subject_id={selected_subject_id}")
        for student in registered_students:
            prefix = f'student_{student.id}_'
            regs = CourseRegistration.query.filter_by(student_id=student.id, subject_id=selected_subject_id).all()
            is_retake = any(r.is_retake for r in regs)
            print(f"[DEBUG] Student {student.student_id} ({student.name}) regs: {[r.is_retake for r in regs]}, is_retake: {is_retake}")
            Mark.query.filter_by(student_id=student.id, subject_id=selected_subject_id).delete()
            if selected_subject.subject_type == 'Theory':
                attendance = safe_float(request.form.get(prefix + 'attendance'))
                continuous_assessment = safe_float(request.form.get(prefix + 'continuous_assessment'))
                part_a = safe_float(request.form.get(prefix + 'part_a'))
                part_b = safe_float(request.form.get(prefix + 'part_b'))
                total_marks = attendance + continuous_assessment + part_a + part_b
                mark = Mark(
                    student_id=student.id,
                    subject_id=selected_subject_id,
                    attendance=attendance,
                    continuous_assessment=continuous_assessment,
                    part_a=part_a,
                    part_b=part_b,
                    is_retake=is_retake
                )
            elif selected_subject.subject_type == 'Dissertation':
                if selected_subject.dissertation_type == 'Type1':
                    supervisor_assessment = safe_float(request.form.get(prefix + 'supervisor_assessment'))
                    proposal_presentation = safe_float(request.form.get(prefix + 'proposal_presentation'))
                    total_marks = supervisor_assessment + proposal_presentation
                    mark = Mark(
                        student_id=student.id,
                        subject_id=selected_subject_id,
                        supervisor_assessment=supervisor_assessment,
                        proposal_presentation=proposal_presentation,
                        is_retake=is_retake
                    )
                else:
                    supervisor_assessment = safe_float(request.form.get(prefix + 'supervisor_assessment'))
                    project_report = safe_float(request.form.get(prefix + 'project_report'))
                    defense = safe_float(request.form.get(prefix + 'defense'))
                    total_marks = supervisor_assessment + project_report + defense
                    mark = Mark(
                        student_id=student.id,
                        subject_id=selected_subject_id,
                        supervisor_assessment=supervisor_assessment,
                        project_report=project_report,
                        defense=defense,
                        is_retake=is_retake
                    )
            grade_point, grade_letter = calculate_grade(total_marks, is_retake)
            mark.total_marks = total_marks
            mark.grade_point = grade_point
            mark.grade_letter = grade_letter
            db.session.add(mark)
        db.session.commit()
        flash('Marks added successfully!', 'success')
        return redirect(url_for('add_marks', session_id=session_id, subject_id=selected_subject_id))
    # Load existing marks for draft/restore
    marks_map = {}
    if selected_subject:
        marks = Mark.query.filter_by(subject_id=selected_subject_id).all()
        for m in marks:
            marks_map[m.student_id] = m
    return render_template('add_marks.html', students=students, subjects=subjects, selected_subject=selected_subject, registered_students=registered_students, marks_map=marks_map, session_id=session_id, is_retake_map=is_retake_map)

@app.route('/generate_result/<int:session_id>')
def generate_result(session_id):
    session = Session.query.get_or_404(session_id)
    students = Student.query.filter_by(session_id=session_id).all()
    subjects = Subject.query.filter_by(session_id=session_id).all()
    
    # Generate course-wise result
    course_result = []
    for subject in subjects:
        subject_data = {
            'code': subject.code,
            'name': subject.name,
            'credit': subject.credit,
            'type': subject.subject_type,
            'marks': []
        }
        
        for student in students:
            mark = Mark.query.filter_by(student_id=student.id, subject_id=subject.id).first()
            if mark:
                remarks = ''
                if mark.is_retake and mark.grade_letter != 'F':
                    remarks = 'Retake'
                subject_data['marks'].append({
                    'student_id': student.student_id,
                    'name': student.name,
                    'total_marks': mark.total_marks,
                    'grade_point': mark.grade_point,
                    'grade_letter': mark.grade_letter,
                    'remarks': remarks
                })
        
        course_result.append(subject_data)
    
    # Generate student-wise result
    student_result = []
    for student in students:
        student_data = {
            'student_id': student.student_id,
            'name': student.name,
            'subjects': []
        }
        
        for subject in subjects:
            mark = Mark.query.filter_by(student_id=student.id, subject_id=subject.id).first()
            if mark:
                remarks = ''
                if mark.is_retake and mark.grade_letter != 'F':
                    remarks = 'Retake'
                student_data['subjects'].append({
                    'code': subject.code,
                    'name': subject.name,
                    'credit': subject.credit,
                    'type': subject.subject_type,
                    'total_marks': mark.total_marks,
                    'grade_point': mark.grade_point,
                    'grade_letter': mark.grade_letter,
                    'remarks': remarks
                })
        
        student_result.append(student_data)
    
    return render_template('generate_result.html',
                         session=session,
                         course_result=course_result,
                         student_result=student_result)

@app.route('/download_course_result/<int:session_id>')
def download_course_result(session_id):
    session = Session.query.get_or_404(session_id)
    students = Student.query.filter_by(session_id=session_id).all()
    subjects = Subject.query.filter_by(session_id=session_id).all()
    
    # Ensure directory exists
    pdf_dir = os.path.join(os.path.dirname(__file__), 'generated_pdfs')
    os.makedirs(pdf_dir, exist_ok=True)
    filename = f'course_result_{session.name}_{session.term}.pdf'
    filepath = os.path.join(pdf_dir, filename)
    doc = SimpleDocTemplate(
        filepath,
        pagesize=letter,
        leftMargin=20,
        rightMargin=20,
        topMargin=30,
        bottomMargin=30
    )
    elements = []
    
    # Add title
    styles = getSampleStyleSheet()
    title = Paragraph(f'Course-wise Result - {session.name} {session.term}', styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 20))
    
    # Create table data
    for subject in subjects:
        elements.append(Paragraph(f'Subject: {subject.code} - {subject.name}', styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        data = [
            [
                'Course No.', 'Course Title', 'Registered Credit Hours', 'Letter Grade', 'Grade Point (GP)',
                'Earned Credit Hours (CH)', 'Earned Credit Points (GP*CH)', 'Remarks'
            ]
        ]
        for student in students:
            mark = Mark.query.filter_by(student_id=student.id, subject_id=subject.id).first()
            if mark:
                remarks = ''
                if mark.is_retake and mark.grade_letter != 'F':
                    remarks = 'Retake'
                data.append([
                    student.student_id,
                    student.name,
                    str(subject.credit),
                    mark.grade_letter if mark else '',
                    str(mark.grade_point) if mark else '',
                    str(subject.credit) if mark and mark.grade_letter != 'F' else '',
                    f'{mark.grade_point * subject.credit:.2f}' if mark and mark.grade_letter != 'F' else '',
                    remarks
                ])
        
        table = Table(data, repeatRows=1, colWidths=None, hAlign='CENTER')
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
        elements.append(Spacer(1, 20))
    
    # Add signature rows to course-wise PDF (single)
    elements.append(Spacer(1, 24))
    sign_table = [
        [
            Paragraph('Signature of the First Tabulator', ParagraphStyle('sign', fontSize=7.5, alignment=1, wordWrap='CJK')),
            Paragraph('Signature of the Second Tabulator', ParagraphStyle('sign', fontSize=7.5, alignment=1, wordWrap='CJK')),
            Paragraph('Signature of the Chairman,<br/>Examination Committee', ParagraphStyle('sign', fontSize=7.5, alignment=1, wordWrap='CJK'))
        ],
        [
            Paragraph('Date:', ParagraphStyle('sign', fontSize=7.5, alignment=1)),
            Paragraph('Date:', ParagraphStyle('sign', fontSize=7.5, alignment=1)),
            Paragraph('Date:', ParagraphStyle('sign', fontSize=7.5, alignment=1))
        ]
    ]
    sign = Table(sign_table, colWidths=[160, 160, 235], hAlign='CENTER')
    sign.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 7.5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(sign)
    doc.build(elements)
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    centered_style = ParagraphStyle(
        name='centered',
        parent=styles['Title'],
        alignment=TA_CENTER
    )
    title = Paragraph('<b>Course-wise Tabulation Sheet</b>', centered_style)
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    # Title (centered)
    elements.append(Paragraph('<b>Khulna University</b>', styles['Title']))
    elements.append(Paragraph('<b>Course-wise Tabulation Sheet</b>', centered_style))
    elements.append(Spacer(1, 12))
    
    # Create table data
    for subject in subjects:
        elements.append(Paragraph(f'Subject: {subject.code} - {subject.name}', styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        data = [
            [
                'Course No.', 'Course Title', 'Registered Credit Hours', 'Letter Grade', 'Grade Point (GP)',
                'Earned Credit Hours (CH)', 'Earned Credit Points (GP*CH)', 'Remarks'
            ]
        ]
        for student in students:
            mark = Mark.query.filter_by(student_id=student.id, subject_id=subject.id).first()
            if mark:
                remarks = ''
                if mark.is_retake and mark.grade_letter != 'F':
                    remarks = 'Retake'
                data.append([
                    student.student_id,
                    student.name,
                    str(subject.credit),
                    mark.grade_letter if mark else '',
                    str(mark.grade_point) if mark else '',
                    str(subject.credit) if mark and mark.grade_letter != 'F' else '',
                    f'{mark.grade_point * subject.credit:.2f}' if mark and mark.grade_letter != 'F' else '',
                    remarks
                ])
        
        table = Table(data, repeatRows=1, colWidths=None, hAlign='CENTER')
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
        elements.append(Spacer(1, 20))
    
    # Add signature rows to course-wise PDF (bulk)
    elements.append(Spacer(1, 24))
    sign_table = [
        [
            Paragraph('Signature of the First Tabulator', ParagraphStyle('sign', fontSize=7.5, alignment=1, wordWrap='CJK')),
            Paragraph('Signature of the Second Tabulator', ParagraphStyle('sign', fontSize=7.5, alignment=1, wordWrap='CJK')),
            Paragraph('Signature of the Chairman,<br/>Examination Committee', ParagraphStyle('sign', fontSize=7.5, alignment=1, wordWrap='CJK'))
        ],
        [
            Paragraph('Date:', ParagraphStyle('sign', fontSize=7.5, alignment=1)),
            Paragraph('Date:', ParagraphStyle('sign', fontSize=7.5, alignment=1)),
            Paragraph('Date:', ParagraphStyle('sign', fontSize=7.5, alignment=1))
        ]
    ]
    sign = Table(sign_table, colWidths=[160, 160, 235], hAlign='CENTER')
    sign.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 7.5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(sign)
    doc.build(elements)
    pdf_buffer.seek(0)
    return send_file(pdf_buffer, as_attachment=True)

@app.route('/download_student_result/<int:session_id>')
def download_student_result(session_id):
    session = Session.query.get_or_404(session_id)
    students = Student.query.filter_by(session_id=session_id).all()
    subjects = Subject.query.filter_by(session_id=session_id).all()
    
    # Ensure directory exists
    pdf_dir = os.path.join(os.path.dirname(__file__), 'generated_pdfs')
    os.makedirs(pdf_dir, exist_ok=True)
    filename = f'student_result_{session.name}_{session.term}.pdf'
    filepath = os.path.join(pdf_dir, filename)
    doc = SimpleDocTemplate(filepath, pagesize=letter)
    elements = []
    
    # Add title
    styles = getSampleStyleSheet()
    title = Paragraph(f'Student-wise Result - {session.name} {session.term}', styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 20))
    
    # Create table data
    for student in students:
        elements.append(Paragraph(f'Student: {student.student_id} - {student.name}', styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        data = [
            [
                'Course No.', 'Course Title', 'Registered Credit Hours', 'Letter Grade', 'Grade Point (GP)',
                'Earned Credit Hours (CH)', 'Earned Credit Points (GP*CH)', 'Remarks'
            ]
        ]
        for subject in subjects:
            mark = Mark.query.filter_by(student_id=student.id, subject_id=subject.id).first()
            if mark:
                remarks = ''
                if mark.is_retake and mark.grade_letter != 'F':
                    remarks = 'Retake'
                data.append([
                    subject.code,
                    subject.name,
                    str(subject.credit),
                    mark.grade_letter if mark else '',
                    str(mark.grade_point) if mark else '',
                    str(subject.credit) if mark and mark.grade_letter != 'F' else '',
                    f'{mark.grade_point * subject.credit:.2f}' if mark and mark.grade_letter != 'F' else '',
                    remarks
                ])
        
        table = Table(data, repeatRows=1, colWidths=None, hAlign='CENTER')
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
        elements.append(Spacer(1, 20))
    
    doc.build(elements)
    filepath = os.path.join(pdf_dir, filename)
    return send_file(filepath, as_attachment=True)

@app.route('/course_registration/<int:session_id>', methods=['GET', 'POST'])
def course_registration(session_id):
    subjects = Subject.query.filter_by(session_id=session_id).all()
    selected_subject_id = request.args.get('subject_id', type=int)
    selected_subject = Subject.query.get(selected_subject_id) if selected_subject_id else None
    students = Student.query.filter_by(session_id=session_id).all() if selected_subject else []
    reg_map = {}
    if selected_subject:
        for reg in CourseRegistration.query.filter_by(subject_id=selected_subject_id).all():
            if reg.is_retake:
                reg_map[(reg.student_id, 'retake')] = True
            else:
                reg_map[(reg.student_id, 'reg')] = True
    if request.method == 'POST' and selected_subject:
        # Clear previous registrations for this subject
        for student in students:
            CourseRegistration.query.filter_by(student_id=student.id, subject_id=selected_subject_id).delete()
        db.session.commit()
        # Add new registrations
        for student in students:
            reg_key = f'student_{student.id}_reg'
            retake_key = f'student_{student.id}_retake'
            is_reg = request.form.get(reg_key) == 'on'
            is_retake = request.form.get(retake_key) == 'on'
            if is_reg:
                reg = CourseRegistration(student_id=student.id, subject_id=selected_subject_id, is_retake=False)
                db.session.add(reg)
            if is_retake:
                reg = CourseRegistration(student_id=student.id, subject_id=selected_subject_id, is_retake=True)
                db.session.add(reg)
        db.session.commit()
        flash('Course registration updated successfully!', 'success')
        return redirect(url_for('course_registration', session_id=session_id, subject_id=selected_subject_id))
    return render_template('course_registration.html', session_id=session_id, subjects=subjects, selected_subject=selected_subject, students=students, reg_map=reg_map)

@app.route('/get_students_for_subject/<int:session_id>/<int:subject_id>')
def get_students_for_subject(session_id, subject_id):
    students = Student.query.filter_by(session_id=session_id).all()
    reg_map = {}
    for reg in CourseRegistration.query.filter_by(subject_id=subject_id).all():
        if reg.is_retake:
            reg_map[(reg.student_id, 'retake')] = True
        else:
            reg_map[(reg.student_id, 'reg')] = True
    data = []
    for student in students:
        data.append({
            'id': student.id,
            'student_id': student.student_id,
            'name': student.name,
            'reg': reg_map.get((student.id, 'reg'), False),
            'retake': reg_map.get((student.id, 'retake'), False)
        })
    return jsonify(data)

@app.route('/view_results/<int:session_id>')
def view_results(session_id):
    return render_template('view_results.html', session_id=session_id)

@app.route('/course_wise_result/<int:session_id>', methods=['GET'])
def course_wise_result(session_id):
    subjects = Subject.query.filter_by(session_id=session_id).all()
    selected_subject_id = request.args.get('subject_id', type=int)
    selected_subject = Subject.query.get(selected_subject_id) if selected_subject_id else None
    students = []
    marks_map = {}
    if selected_subject:
        regs = CourseRegistration.query.filter_by(subject_id=selected_subject_id).all()
        student_ids = [reg.student_id for reg in regs]
        students = Student.query.filter(Student.id.in_(student_ids)).all()
        marks = Mark.query.filter(Mark.subject_id==selected_subject_id, Mark.student_id.in_(student_ids)).all()
        for m in marks:
            marks_map[m.student_id] = m
    return render_template('course_wise_result.html', session_id=session_id, subjects=subjects, selected_subject=selected_subject, students=students, marks_map=marks_map)

@app.route('/student_wise_result/<int:session_id>', methods=['GET'])
def student_wise_result(session_id):
    students = Student.query.filter_by(session_id=session_id).all()
    selected_student_id = request.args.get('student_id', type=int)
    selected_student = Student.query.get(selected_student_id) if selected_student_id else None
    subjects = []
    marks_map = {}
    if selected_student:
        regs = CourseRegistration.query.filter_by(student_id=selected_student_id).all()
        subject_ids = [reg.subject_id for reg in regs]
        subjects = Subject.query.filter(Subject.id.in_(subject_ids)).all()
        marks = Mark.query.filter(Mark.student_id==selected_student_id, Mark.subject_id.in_(subject_ids)).all()
        for m in marks:
            marks_map[m.subject_id] = m
    return render_template('student_wise_result.html', session_id=session_id, students=students, selected_student=selected_student, subjects=subjects, marks_map=marks_map)

@app.route('/download_single_student_result_pdf/<int:session_id>/<int:student_id>')
def download_single_student_result_pdf(session_id, student_id):
    session = Session.query.get_or_404(session_id)
    student = Student.query.get_or_404(student_id)
    regs = CourseRegistration.query.filter_by(student_id=student_id).all()
    subject_ids = [reg.subject_id for reg in regs]
    subjects = Subject.query.filter(Subject.id.in_(subject_ids)).all()
    marks = Mark.query.filter(Mark.student_id==student_id, Mark.subject_id.in_(subject_ids)).all()
    marks_map = {m.subject_id: m for m in marks}

    # Calculate totals
    rch = sum(s.credit for s in subjects)
    tch = sum(s.credit for s in subjects if marks_map.get(s.id) and marks_map[s.id].grade_letter != 'F')
    tcp = sum((marks_map[s.id].grade_point * s.credit) for s in subjects if marks_map.get(s.id))
    tgpa = tcp / rch if rch > 0 else 0

    # PDF generation
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()
    header_style = ParagraphStyle('header', fontSize=8, alignment=1, textColor=colors.whitesmoke, fontName='Helvetica-Bold')
    elements.append(Paragraph('<b>Khulna University</b>', styles['Title']))
    elements.append(Paragraph('Student-wise Tabulation Sheet', styles['Heading2']))
    elements.append(Spacer(1, 12))
    # Student info
    info_table = [
        [
            Paragraph(f'<b>Year:</b> LL.M.', styles['Normal']),
            Paragraph(f'<b>Term:</b> {session.term}', styles['Normal']),
            Paragraph(f'<b>Session:</b> {session.name}', styles['Normal'])
        ],
        [
            Paragraph(f'<b>Student No.:</b> {student.student_id}', styles['Normal']),
            Paragraph(f'<b>Name of Student:</b> {student.name}', styles['Normal']),
            Paragraph(f'<b>School:</b> Law', styles['Normal'])
        ],
        [
            Paragraph(f'<b>Discipline:</b> Law', styles['Normal']),
            '',
            ''
        ]
    ]
    elements.append(Table(info_table, colWidths=[180, 180, 180]))
    elements.append(Spacer(1, 12))
    # Table header (FIXED)
    header_cells = [
        Paragraph('Course No.', header_style),
        Paragraph('Course Title', header_style),
        Paragraph('Registered<br/>Credit Hours', header_style),
        Paragraph('Letter<br/>Grade', header_style),
        Paragraph('Grade<br/>Point (GP)', header_style),
        Paragraph('Earned<br/>Credit Hours (CH)', header_style),
        Paragraph('Earned Credit<br/>Points (GP*CH)', header_style),
        Paragraph('Remarks', header_style)
    ]
    colWidths = [80, 150, 60, 45, 45, 60, 65, 55]
    data = [header_cells]
    for subject in subjects:
        mark = marks_map.get(subject.id)
        remarks = ''
        if mark and mark.is_retake and mark.grade_letter != 'F':
            remarks = 'Retake'
        data.append([
            Paragraph(subject.code, styles['Normal']),
            Paragraph(subject.name, styles['Normal']),
            str(subject.credit),
            mark.grade_letter if mark else '',
            str(mark.grade_point) if mark else '',
            str(subject.credit) if mark and mark.grade_letter != 'F' else '',
            f'{mark.grade_point * subject.credit:.2f}' if mark and mark.grade_letter != 'F' else '',
            remarks
        ])
    table = Table(data, repeatRows=1, colWidths=colWidths, hAlign='CENTER')
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(table)
    elements.append(Spacer(1, 16))
    # Restore signature rows to student-wise PDF (single)
    elements.append(Paragraph('<b>Term Assessment</b>', styles['Normal']))
    elements.append(Paragraph(f'Total Earned Credit Hours in this Term (TCH) = {tch}', styles['Normal']))
    elements.append(Paragraph(f'Total Registered Credit Hours in this Term (RCH) = {rch}', styles['Normal']))
    elements.append(Paragraph(f'Total Earned Credit Points in this Term (TCP) = {tcp:.2f}', styles['Normal']))
    elements.append(Paragraph(f'TGPA = TCP/RCH = {tgpa:.2f}' if rch > 0 else 'TGPA = TCP/RCH = N/A', styles['Normal']))
    elements.append(Spacer(1, 24))
    sign_table = [
        [
            Paragraph('Signature of the First Tabulator', ParagraphStyle('sign', fontSize=7.5, alignment=1, wordWrap='CJK')),
            Paragraph('Signature of the Second Tabulator', ParagraphStyle('sign', fontSize=7.5, alignment=1, wordWrap='CJK')),
            Paragraph('Signature of the Chairman,<br/>Examination Committee', ParagraphStyle('sign', fontSize=7.5, alignment=1, wordWrap='CJK'))
        ],
        [
            Paragraph('Date:', ParagraphStyle('sign', fontSize=7.5, alignment=1)),
            Paragraph('Date:', ParagraphStyle('sign', fontSize=7.5, alignment=1)),
            Paragraph('Date:', ParagraphStyle('sign', fontSize=7.5, alignment=1))
        ]
    ]
    sign = Table(sign_table, colWidths=[160, 160, 235], hAlign='CENTER')
    sign.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 7.5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(sign)
    doc.build(elements)
    pdf_buffer.seek(0)
    return send_file(pdf_buffer, as_attachment=True, download_name=f'student_result_{student.student_id}.pdf', mimetype='application/pdf')

@app.route('/download_bulk_student_result_pdf/<int:session_id>')
def download_bulk_student_result_pdf(session_id):
    session = Session.query.get_or_404(session_id)
    students = Student.query.filter_by(session_id=session_id).all()
    # Prepare in-memory zip
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w') as zipf:
        for student in students:
            regs = CourseRegistration.query.filter_by(student_id=student.id).all()
            subject_ids = [reg.subject_id for reg in regs]
            subjects = Subject.query.filter(Subject.id.in_(subject_ids)).all()
            marks = Mark.query.filter(Mark.student_id==student.id, Mark.subject_id.in_(subject_ids)).all()
            marks_map = {m.subject_id: m for m in marks}
            rch = sum(s.credit for s in subjects)
            tch = sum(s.credit for s in subjects if marks_map.get(s.id) and marks_map[s.id].grade_letter != 'F')
            tcp = sum((marks_map[s.id].grade_point * s.credit) for s in subjects if marks_map.get(s.id))
            tgpa = tcp / rch if rch > 0 else 0
            pdf_buffer = io.BytesIO()
            doc = SimpleDocTemplate(pdf_buffer, pagesize=landscape(A4))
            elements = []
            styles = getSampleStyleSheet()
            header_style = ParagraphStyle('header', fontSize=8, alignment=1, textColor=colors.whitesmoke, fontName='Helvetica-Bold')
            elements.append(Paragraph('<b>Khulna University</b>', styles['Title']))
            elements.append(Paragraph('Student-wise Tabulation Sheet', styles['Heading2']))
            elements.append(Spacer(1, 12))
            info_table = [
                [
                    Paragraph(f'<b>Year:</b> LL.M.', styles['Normal']),
                    Paragraph(f'<b>Term:</b> {session.term}', styles['Normal']),
                    Paragraph(f'<b>Session:</b> {session.name}', styles['Normal'])
                ],
                [
                    Paragraph(f'<b>Student No.:</b> {student.student_id}', styles['Normal']),
                    Paragraph(f'<b>Name of Student:</b> {student.name}', styles['Normal']),
                    Paragraph(f'<b>School:</b> Law', styles['Normal'])
                ],
                [
                    Paragraph(f'<b>Discipline:</b> Law', styles['Normal']),
                    '',
                    ''
                ]
            ]
            elements.append(Table(info_table, colWidths=[180, 180, 180]))
            elements.append(Spacer(1, 12))
            # Table header
            header_cells = [
                Paragraph('Course No.', header_style),
                Paragraph('Course Title', header_style),
                Paragraph('Registered<br/>Credit Hours', header_style),
                Paragraph('Letter<br/>Grade', header_style),
                Paragraph('Grade<br/>Point (GP)', header_style),
                Paragraph('Earned<br/>Credit Hours (CH)', header_style),
                Paragraph('Earned Credit<br/>Points (GP*CH)', header_style),
                Paragraph('Remarks', header_style)
            ]
            data = [header_cells]
            for subject in subjects:
                mark = marks_map.get(subject.id)
                remarks = ''
                if mark and mark.is_retake and mark.grade_letter != 'F':
                    remarks = 'Retake'
                data.append([
                    Paragraph(subject.code, styles['Normal']),
                    Paragraph(subject.name, styles['Normal']),
                    str(subject.credit),
                    mark.grade_letter if mark else '',
                    str(mark.grade_point) if mark else '',
                    str(subject.credit) if mark and mark.grade_letter != 'F' else '',
                    f'{mark.grade_point * subject.credit:.2f}' if mark and mark.grade_letter != 'F' else '',
                    remarks
                ])
            table = Table(data, repeatRows=1, colWidths=None, hAlign='CENTER')
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(table)
            elements.append(Spacer(1, 16))
            # Restore signature rows to student-wise PDF (bulk)
            elements.append(Paragraph('<b>Term Assessment</b>', styles['Normal']))
            elements.append(Paragraph(f'Total Earned Credit Hours in this Term (TCH) = {tch}', styles['Normal']))
            elements.append(Paragraph(f'Total Registered Credit Hours in this Term (RCH) = {rch}', styles['Normal']))
            elements.append(Paragraph(f'Total Earned Credit Points in this Term (TCP) = {tcp:.2f}', styles['Normal']))
            elements.append(Paragraph(f'TGPA = TCP/RCH = {tgpa:.2f}' if rch > 0 else 'TGPA = TCP/RCH = N/A', styles['Normal']))
            elements.append(Spacer(1, 24))
            sign_table = [
                [
                    Paragraph('Signature of the First Tabulator', ParagraphStyle('sign', fontSize=7.5, alignment=1, wordWrap='CJK')),
                    Paragraph('Signature of the Second Tabulator', ParagraphStyle('sign', fontSize=7.5, alignment=1, wordWrap='CJK')),
                    Paragraph('Signature of the Chairman,<br/>Examination Committee', ParagraphStyle('sign', fontSize=7.5, alignment=1, wordWrap='CJK'))
                ],
                [
                    Paragraph('Date:', ParagraphStyle('sign', fontSize=7.5, alignment=1)),
                    Paragraph('Date:', ParagraphStyle('sign', fontSize=7.5, alignment=1)),
                    Paragraph('Date:', ParagraphStyle('sign', fontSize=7.5, alignment=1))
                ]
            ]
            sign = Table(sign_table, colWidths=[160, 160, 235], hAlign='CENTER')
            sign.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTSIZE', (0, 0), (-1, -1), 7.5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            elements.append(sign)
            doc.build(elements)
            pdf_buffer.seek(0)
            zipf.writestr(f'student_result_{student.student_id}.pdf', pdf_buffer.read())
    zip_buffer.seek(0)
    return send_file(zip_buffer, as_attachment=True, download_name=f'student_results_{session.name}.zip', mimetype='application/zip')

@app.route('/download_single_course_result_pdf/<int:session_id>/<int:subject_id>')
def download_single_course_result_pdf(session_id, subject_id):
    session = Session.query.get_or_404(session_id)
    subject = Subject.query.get_or_404(subject_id)
    regs = CourseRegistration.query.filter_by(subject_id=subject_id).all()
    student_ids = [reg.student_id for reg in regs]
    students = Student.query.filter(Student.id.in_(student_ids)).all()
    marks = Mark.query.filter(Mark.subject_id==subject_id, Mark.student_id.in_(student_ids)).all()
    marks_map = {m.student_id: m for m in marks}

    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=A4, leftMargin=10, rightMargin=10, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()
    header_style = ParagraphStyle('header', fontSize=8, alignment=1, textColor=colors.whitesmoke, fontName='Helvetica-Bold')
    elements.append(Paragraph('<b>Khulna University</b>', styles['Title']))
    elements.append(Paragraph('Course-wise Tabulation Sheet', styles['Heading2']))
    elements.append(Spacer(1, 12))
    info_table = [
        [
            Paragraph(f'<b>Year:</b> LL.M.', styles['Normal']),
            Paragraph(f'<b>Term:</b> {subject.session.term}', styles['Normal']),
            Paragraph(f'<b>Session:</b> {subject.session.name}', styles['Normal'])
        ],
        [
            Paragraph(f'<b>Discipline:</b> Law', styles['Normal']),
            Paragraph(f'<b>School:</b> Law', styles['Normal']),
            ''
        ],
        [
            Paragraph(f'<b>Course No.:</b> {subject.code}', styles['Normal']),
            Paragraph(f'<b>CH:</b> {subject.credit}', styles['Normal']),
            ''
        ],
        [
            Paragraph(f'<b>Course Title:</b> {subject.name}', styles['Normal']),
            '',
            ''
        ]
    ]
    elements.append(Table(info_table, colWidths=[180, 180, 180]))
    elements.append(Spacer(1, 12))
    # Table header
    if subject.subject_type == 'Theory':
        header_cells = [
            Paragraph('Student No.', header_style),
            Paragraph('Attendance<br/>(10)', header_style),
            Paragraph('Continuous<br/>Assessment (40)', header_style),
            Paragraph('Section A<br/>(25)', header_style),
            Paragraph('Section B<br/>(25)', header_style),
            Paragraph('Total Marks<br/>(100)', header_style),
            Paragraph('Grade<br/>Point', header_style),
            Paragraph('Grade<br/>Letter', header_style),
            Paragraph('Remarks', header_style)
        ]
        colWidths = [50, 45, 70, 50, 50, 50, 45, 45, 55]
    elif subject.subject_type == 'Dissertation' and subject.dissertation_type == 'Type1':
        header_cells = [
            Paragraph('Student No.', header_style),
            Paragraph('Supervisor<br/>Assessment (30)', header_style),
            Paragraph('Proposal<br/>Presentation (70)', header_style),
            Paragraph('Total Marks<br/>(100)', header_style),
            Paragraph('Grade<br/>Point', header_style),
            Paragraph('Grade<br/>Letter', header_style),
            Paragraph('Remarks', header_style)
        ]
        colWidths = [50, 90, 90, 50, 45, 45, 55]
    elif subject.subject_type == 'Dissertation' and subject.dissertation_type == 'Type2':
        header_cells = [
            Paragraph('Student No.', header_style),
            Paragraph('Supervisor<br/>Assessment (20)', header_style),
            Paragraph('Project<br/>Report (50)', header_style),
            Paragraph('Defense<br/>(30)', header_style),
            Paragraph('Total Marks<br/>(100)', header_style),
            Paragraph('Grade<br/>Point', header_style),
            Paragraph('Grade<br/>Letter', header_style),
            Paragraph('Remarks', header_style)
        ]
        colWidths = [50, 60, 90, 60, 50, 45, 45, 55]
    else:
        colWidths = [55] * len(header_cells)
    data = [header_cells]
    for student in students:
        mark = marks_map.get(student.id)
        remarks = ''
        if mark and mark.is_retake and mark.grade_letter != 'F':
            remarks = 'Retake'
        row = [student.student_id]
        if subject.subject_type == 'Theory':
            row += [
                mark.attendance if mark else '',
                mark.continuous_assessment if mark else '',
                mark.part_a if mark else '',
                mark.part_b if mark else ''
            ]
        elif subject.subject_type == 'Dissertation' and subject.dissertation_type == 'Type1':
            row += [
                mark.supervisor_assessment if mark else '',
                mark.proposal_presentation if mark else ''
            ]
        elif subject.subject_type == 'Dissertation' and subject.dissertation_type == 'Type2':
            row += [
                mark.supervisor_assessment if mark else '',
                mark.project_report if mark else '',
                mark.defense if mark else ''
            ]
        row += [
            mark.total_marks if mark else '',
            mark.grade_point if mark else '',
            mark.grade_letter if mark else '',
            remarks
        ]
        data.append(row)
    table = Table(data, repeatRows=1, colWidths=colWidths, hAlign='CENTER')
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(table)
    elements.append(Spacer(1, 24))
    # Add signature rows to course-wise PDF (single)
    sign_table = [
        [
            Paragraph('Signature of the First Tabulator', ParagraphStyle('sign', fontSize=7.5, alignment=1, wordWrap='CJK')),
            Paragraph('Signature of the Second Tabulator', ParagraphStyle('sign', fontSize=7.5, alignment=1, wordWrap='CJK')),
            Paragraph('Signature of the Chairman,<br/>Examination Committee', ParagraphStyle('sign', fontSize=7.5, alignment=1, wordWrap='CJK'))
        ],
        [
            Paragraph('Date:', ParagraphStyle('sign', fontSize=7.5, alignment=1)),
            Paragraph('Date:', ParagraphStyle('sign', fontSize=7.5, alignment=1)),
            Paragraph('Date:', ParagraphStyle('sign', fontSize=7.5, alignment=1))
        ]
    ]
    sign = Table(sign_table, colWidths=[160, 160, 235], hAlign='CENTER')
    sign.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 7.5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(sign)
    doc.build(elements)
    pdf_buffer.seek(0)
    return send_file(pdf_buffer, as_attachment=True, download_name=f'course_result_{subject.code}.pdf', mimetype='application/pdf')

@app.route('/download_bulk_course_result_pdf/<int:session_id>')
def download_bulk_course_result_pdf(session_id):
    session = Session.query.get_or_404(session_id)
    subjects = Subject.query.filter_by(session_id=session_id).all()
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w') as zipf:
        for subject in subjects:
            regs = CourseRegistration.query.filter_by(subject_id=subject.id).all()
            student_ids = [reg.student_id for reg in regs]
            students = Student.query.filter(Student.id.in_(student_ids)).all()
            marks = Mark.query.filter(Mark.subject_id==subject.id, Mark.student_id.in_(student_ids)).all()
            marks_map = {m.student_id: m for m in marks}
            pdf_buffer = io.BytesIO()
            doc = SimpleDocTemplate(pdf_buffer, pagesize=A4, leftMargin=10, rightMargin=10, topMargin=30, bottomMargin=30)
            elements = []
            styles = getSampleStyleSheet()
            header_style = ParagraphStyle('header', fontSize=8, alignment=1, textColor=colors.whitesmoke, fontName='Helvetica-Bold')
            elements.append(Paragraph('<b>Khulna University</b>', styles['Title']))
            elements.append(Paragraph('Course-wise Tabulation Sheet', styles['Heading2']))
            elements.append(Spacer(1, 12))
            info_table = [
                [
                    Paragraph(f'<b>Year:</b> LL.M.', styles['Normal']),
                    Paragraph(f'<b>Term:</b> {subject.session.term}', styles['Normal']),
                    Paragraph(f'<b>Session:</b> {subject.session.name}', styles['Normal'])
                ],
                [
                    Paragraph(f'<b>Discipline:</b> Law', styles['Normal']),
                    Paragraph(f'<b>School:</b> Law', styles['Normal']),
                    ''
                ],
                [
                    Paragraph(f'<b>Course No.:</b> {subject.code}', styles['Normal']),
                    Paragraph(f'<b>CH:</b> {subject.credit}', styles['Normal']),
                    ''
                ],
                [
                    Paragraph(f'<b>Course Title:</b> {subject.name}', styles['Normal']),
                    '',
                    ''
                ]
            ]
            elements.append(Table(info_table, colWidths=[180, 180, 180]))
            elements.append(Spacer(1, 12))
            data = [
                [
                    'Course No.', 'Course Title', 'Registered Credit Hours', 'Letter Grade', 'Grade Point (GP)',
                    'Earned Credit Hours (CH)', 'Earned Credit Points (GP*CH)', 'Remarks'
                ]
            ]
            if subject.subject_type == 'Theory':
                header_cells = [
                    Paragraph('Student No.', header_style),
                    Paragraph('Attendance<br/>(10)', header_style),
                    Paragraph('Continuous<br/>Assessment (40)', header_style),
                    Paragraph('Section A<br/>(25)', header_style),
                    Paragraph('Section B<br/>(25)', header_style),
                    Paragraph('Total Marks<br/>(100)', header_style),
                    Paragraph('Grade<br/>Point', header_style),
                    Paragraph('Grade<br/>Letter', header_style),
                    Paragraph('Remarks', header_style)
                ]
                colWidths = [50, 45, 70, 50, 50, 50, 45, 45, 55]
            elif subject.subject_type == 'Dissertation' and subject.dissertation_type == 'Type1':
                header_cells = [
                    Paragraph('Student No.', header_style),
                    Paragraph('Supervisor<br/>Assessment (30)', header_style),
                    Paragraph('Proposal<br/>Presentation (70)', header_style),
                    Paragraph('Total Marks<br/>(100)', header_style),
                    Paragraph('Grade<br/>Point', header_style),
                    Paragraph('Grade<br/>Letter', header_style),
                    Paragraph('Remarks', header_style)
                ]
                colWidths = [50, 90, 90, 50, 45, 45, 55]
            elif subject.subject_type == 'Dissertation' and subject.dissertation_type == 'Type2':
                header_cells = [
                    Paragraph('Student No.', header_style),
                    Paragraph('Supervisor<br/>Assessment (20)', header_style),
                    Paragraph('Project<br/>Report (50)', header_style),
                    Paragraph('Defense<br/>(30)', header_style),
                    Paragraph('Total Marks<br/>(100)', header_style),
                    Paragraph('Grade<br/>Point', header_style),
                    Paragraph('Grade<br/>Letter', header_style),
                    Paragraph('Remarks', header_style)
                ]
                colWidths = [50, 60, 90, 60, 50, 45, 45, 55]
            else:
                header_cells = [Paragraph(h, header_style) for h in data[0]]
                colWidths = [55] * len(header_cells)
            data = [header_cells]
            for student in students:
                mark = marks_map.get(student.id)
                remarks = ''
                if mark and mark.is_retake and mark.grade_letter != 'F':
                    remarks = 'Retake'
                row = [student.student_id]
                if subject.subject_type == 'Theory':
                    row += [
                        mark.attendance if mark else '',
                        mark.continuous_assessment if mark else '',
                        mark.part_a if mark else '',
                        mark.part_b if mark else ''
                    ]
                elif subject.subject_type == 'Dissertation' and subject.dissertation_type == 'Type1':
                    row += [
                        mark.supervisor_assessment if mark else '',
                        mark.proposal_presentation if mark else ''
                    ]
                elif subject.subject_type == 'Dissertation' and subject.dissertation_type == 'Type2':
                    row += [
                        mark.supervisor_assessment if mark else '',
                        mark.project_report if mark else '',
                        mark.defense if mark else ''
                    ]
                row += [
                    mark.total_marks if mark else '',
                    mark.grade_point if mark else '',
                    mark.grade_letter if mark else '',
                    remarks
                ]
                data.append(row)
            table = Table(data, repeatRows=1, colWidths=colWidths, hAlign='CENTER')
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 7),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(table)
            elements.append(Spacer(1, 24))
            # Add signature rows to course-wise PDF (bulk)
            sign_table = [
                [
                    Paragraph('Signature of the First Tabulator', ParagraphStyle('sign', fontSize=7.5, alignment=1, wordWrap='CJK')),
                    Paragraph('Signature of the Second Tabulator', ParagraphStyle('sign', fontSize=7.5, alignment=1, wordWrap='CJK')),
                    Paragraph('Signature of the Chairman,<br/>Examination Committee', ParagraphStyle('sign', fontSize=7.5, alignment=1, wordWrap='CJK'))
                ],
                [
                    Paragraph('Date:', ParagraphStyle('sign', fontSize=7.5, alignment=1)),
                    Paragraph('Date:', ParagraphStyle('sign', fontSize=7.5, alignment=1)),
                    Paragraph('Date:', ParagraphStyle('sign', fontSize=7.5, alignment=1))
                ]
            ]
            sign = Table(sign_table, colWidths=[160, 160, 235], hAlign='CENTER')
            sign.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTSIZE', (0, 0), (-1, -1), 7.5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            elements.append(sign)
            doc.build(elements)
            pdf_buffer.seek(0)
            zipf.writestr(f'course_result_{subject.code}.pdf', pdf_buffer.read())
    zip_buffer.seek(0)
    return send_file(zip_buffer, as_attachment=True, download_name=f'course_results_{session.name}.zip', mimetype='application/zip')

@app.route('/delete_session/<int:session_id>', methods=['POST'])
def delete_session(session_id):
    session = Session.query.get_or_404(session_id)
    # Cascade delete: students, subjects, marks, course registrations
    students = Student.query.filter_by(session_id=session_id).all()
    for student in students:
        Mark.query.filter_by(student_id=student.id).delete()
        CourseRegistration.query.filter_by(student_id=student.id).delete()
        db.session.delete(student)
    subjects = Subject.query.filter_by(session_id=session_id).all()
    for subject in subjects:
        Mark.query.filter_by(subject_id=subject.id).delete()
        CourseRegistration.query.filter_by(subject_id=subject.id).delete()
        db.session.delete(subject)
    db.session.delete(session)
    db.session.commit()
    flash('Session and all related data deleted successfully!', 'success')
    return redirect(url_for('index'))

# --- Authentication Decorator ---
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# --- Registration Route ---
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        email = request.form['email']
        role = 'student'  # Default role
        
        if User.query.filter_by(username=username).first():
            flash('Username already exists')
            return redirect(url_for('register'))
            
        if User.query.filter_by(email=email).first():
            flash('Email already exists')
            return redirect(url_for('register'))
            
        user = User(username=username, email=email, role=role)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        
        flash('Registration successful')
        return redirect(url_for('login'))
        
    return render_template('register.html')

# --- Login Route ---
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password_hash, password):
            session['user_id'] = user.id
            session['username'] = user.username
            flash('Logged in successfully!', 'success')
            return redirect(url_for('index'))
        else:
            flash('Invalid username or password.', 'danger')
    return render_template('login.html')

# --- Logout Route ---
@app.route('/logout')
def logout():
    session.clear()
    flash('Logged out successfully.', 'success')
    return redirect(url_for('login'))

# --- Forgot Password Route (simple version) ---
@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email = request.form['email']
        user = User.query.filter_by(email=email).first()
        if user:
            new_password = request.form['new_password']
            user.password_hash = generate_password_hash(new_password)
            db.session.commit()
            flash('Password reset successful. Please log in.', 'success')
            return redirect(url_for('login'))
        else:
            flash('No user found with that email.', 'danger')
    return render_template('forgot_password.html')

# --- Protect all routes except login/register/forgot_password/static ---
@app.before_request
def require_login():
    allowed_routes = ['login', 'register', 'forgot_password', 'static']
    endpoint = request.endpoint
    if endpoint is None:
        return  # No endpoint, do nothing
    if endpoint not in allowed_routes and not endpoint.startswith('static'):
        if 'user_id' not in session:
            return redirect(url_for('login'))

def safe_float(val):
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0

if __name__ == '__main__':
    app.run(debug=True)
